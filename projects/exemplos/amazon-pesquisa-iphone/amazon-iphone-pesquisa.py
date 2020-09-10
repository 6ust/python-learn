#REALIZAÇÃO DE SCRAPING COM PYTHON UTILIZANDO A BIBLIOTECA [BEAUTIFULSOUP4]
# 
# Antes de Executar o Script
# Rode o Seguinte comando no terminal
#
# pip install beautifulsoup4
# pip install --upgrade edx-dl
# pip install openpyxl

from urllib.request import urlopen
from urllib.error import HTTPError
from urllib.error import URLError
from bs4 import BeautifulSoup


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment, Protection, Font

import re

#URL COM A PESQUISA IPHONE
urlPass = "https://www.amazon.com.br/s?k=iphone&__mk_pt_BR=%C3%85M%C3%85%C5%BD%C3%95%C3%91&ref=nb_sb_noss_2"


names = []


# INSERÇÃO DE DADOS EM PLANILHA
file_excel = Workbook()


#COMANDO UTILIZAR PLANILHA PADRAO
plan1 = file_excel.active


#RENOMEANDO A PLANILHA ATIVA
plan1.title = "Iphone"


# TRATAMENTO DE ERROS DE CONEXÃO
try:
    html = urlopen(urlPass)
except HTTPError as e:
    print(e)
except URLError:
    print("Server down or incorrect domain")
else:

	# PESQUISA DE TAG CLASSE
    res = BeautifulSoup(html.read(),"html.parser")
    tagNames = res.findAll("span", ["a-size-base-plus a-color-base a-text-normal" , "a-offscreen"])


for tt in tagNames:
	names.append(tt.get_text())

for tt in names:
	print(tt)
	break



lineExcel = 0



for i in range(len(names)):
	#EXIBIÇÃO DE NOMES DOS PRODUTOS 
	if re.search(r".[a-z]|[a-z]*PHONE", names[i]) != None:
		plan1['B' + str(lineExcel + 3)] = names[i]

	# EXIBIÇÃO DE PREÇO DOS PRODUTOS 
	if re.search(r"R[$].", names[i]) != None:
		plan1['C' + str(lineExcel + 2)] = names[i]

	lineExcel += 1

# CONFIGURAÇÃO DA FONT DO HEADER
plan1['B2'].font = Font(name='Amiri', size=18, bold=True, color='0000FF')
plan1['C2'].font = Font(name='Amiri', size=18, bold=True, color='0000FF')


# CONFIGURAÇÃO DE HEADER DA PLANILHA
plan1['B2'] = "Nome"
plan1['C2'] = "Preço"

# CONFIGURAÇÃO DA LARGURA DA COLUNA
plan1.column_dimensions['B'].width = 80;
plan1.column_dimensions['C'].width = 30;

# CONFIGURAÇÃO DE ALINHAMENTO DAS CELULAS TITULO
plan1['B2'].alignment = Alignment(horizontal="center", vertical="center")
plan1['C2'].alignment = Alignment(horizontal="center", vertical="center")



# APLICAÇÃO DO FILTRO
plan1.auto_filter.ref = "B2:C400"
plan1.auto_filter.add_sort_condition("B2:B400")


# pn.font = Font(name='Amiri', size=14)

#SALVAR PLANILHA
file_excel.save("amazon-iphone.xlsx")
