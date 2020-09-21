# INSTALAR AS BIBLIOTECAS ABAIXO
# pip install openpyxl
# pip install beautifulsoup4
# pip install --upgrade edx-dl


import os.path

# BIBLIOTECA PARA UTILIZAÇÃO DE PLANILHAS
from openpyxl import Workbook, load_workbook
from openpyxl.utils import FORMULAE, get_column_letter
from openpyxl.cell import Cell
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment, Protection, Font

# BIBLIOTECA PARA USO DE DATA E HORA
from datetime import date, datetime
from time import gmtime, strftime 

# BIBLIOTECA PARA CONFIGURAÇÃO DE HORA POR REGIAO
import pytz

# BIBLIOTECA PARA UTILIZAÇÃO DO BEUTIFULSOUP4
from bs4 import BeautifulSoup

# BIBLIOTECA PARA UTILIZAÇÃO DE REQUESTS
import requests



# OBTER A HORA UTC PADRÃO
UTC = pytz.utc 

# OBTERÁ O FUSO HORÁRIO
# DA LOCALIZAÇÃO ESPECIFICADA
IST = pytz.timezone('Brazil/East') 


# PRINT O DATE E TIME NO
# FORMATO ESPECIFICADO 
datetime_ist = datetime.now(IST) 

# MONTAGEM DO NOME DA PLANILHA
path = "cotacao-" + str(date.today()) + "-time-save-" + str(datetime_ist.strftime('%H-%M-%S')) + ".xlsx"

# CRIAÇÃO DA PLANILHA
file  = Workbook()

planCotacao = file.active 

# TITULOS DA PLANILHA COTAÇÃO
planCotacao["B2"] = "Nome"
planCotacao["C2"] = "Data de Consulta"
planCotacao["D2"] = "Preço"


# --------------------------------------------------
# FORMATAÇÃO DOS TITULOS DA PLANILHA COTAÇÃO
# --------------------------------------------------

# TAMANHO
planCotacao.column_dimensions['B'].width = 30
planCotacao.column_dimensions['C'].width = 40
planCotacao.column_dimensions['D'].width = 30

# FONTES
planCotacao['B2'].font = Font(name='Bahnschrift Light Condensed', size=22, bold=True, color='781DA1')
planCotacao['C2'].font = Font(name='Bahnschrift Light Condensed', size=22, bold=True, color='781DA1')
planCotacao['D2'].font = Font(name='Bahnschrift Light Condensed', size=22, bold=True, color='781DA1')

# ALINHAMENTO DE TEXTO DENTRO DA CELULA
planCotacao['B2'].alignment = Alignment(horizontal="center", vertical="center")
planCotacao['C2'].alignment = Alignment(horizontal="center", vertical="center")
planCotacao['D2'].alignment = Alignment(horizontal="center", vertical="center")

# --------------------------------------------------


prod1 = "https://www.casasbahia.com.br/Eletroportateis/FerrodePassar/FerroaVapor/ferro-a-vapor-mondial-f32-com-spray-branco-azul-4646594.html?utm_medium=Cpc&utm_source=GP_PLA&IdSku=4646594&idLojista=10037&utm_campaign=elpo_smart-shopping&gclid=Cj0KCQjww_f2BRC-ARIsAP3zarGj_A5FLX9pSzTGSHTAMRcGAlzZojGVu4w7QeCQYFsjpGg8BQBF1Z8aAu7zEALw_wcB"
prod2 = "https://www.casasbahia.com.br/UtilidadesDomesticas/lavanderiaebanheiro/tabuadepassar/tabua-de-passar-tramontina-classic-dobravel-31415.html?utm_medium=Cpc&utm_source=GP_PLA&IdSku=31415&idLojista=10037&utm_campaign=utld_smart-shopping&gclid=Cj0KCQjww_f2BRC-ARIsAP3zarEubr2OGqZ01oNGLkSAG6EUfN68SQD7hs8MS95mlVi1sfkITs6bOd8aAoyCEALw_wcB"
prod3 = "https://www.madeiramadeira.com.br/cozinha-compacta-11-portas-4-pecas-eva-yescasa-566863.html?origem=pla-566863&utm_source=google&utm_medium=cpc&utm_content=cozinha-compacta&utm_term=566863&gclid=Cj0KCQjww_f2BRC-ARIsAP3zarHkniu-bqXUBs_IlVMF3DowQl77HX1D8xi87GIQ8gAvWXl25a2PMiIaAvLyEALw_wcB"


# PESQUISA POR URL SOBRE UM PRODUTO
hProd1 = requests.get(prod1).content
hProd2 = requests.get(prod2).content
hProd3 = requests.get(prod3).content



# RETORNO DE TODO CODIGO HTML DO SITE
prod1 = BeautifulSoup(hProd1, "html.parser")
prod2 = BeautifulSoup(hProd2, "html.parser")
prod3 = BeautifulSoup(hProd3, "html.parser")


# ------------------------------
# -- NOMES E VALORES -----------
# ------------------------------
# PRODUTO 1
name1 = prod1.findAll("h1", class_="fn")
prc1 = prod1.findAll("i", class_="sale price")

# PRODUTO 2
name2 = prod2.findAll("h1", class_="fn")
prc2 = prod2.findAll("i", class_="sale price")

# PRODUTO 3
name3 = prod3.findAll("h1", class_="product-title")
prc3 = prod3.findAll("span", class_="typography typography--font-color-default typography--font-size-bigger")

# ------------------------------


# ------------------------------
# PREENCHIMENTO DO PRODUTOs 
# ------------------------------

# PRODUTO 1
planCotacao['B3'] = name1[0].getText()
planCotacao['C3'] = date.today()
planCotacao['D3'] = str("R$") + prc1[0].getText()

# PRODUTO 2
planCotacao['B4'] = name2[0].getText()
planCotacao['C4'] = date.today()
planCotacao['D4'] = str("R$") + prc2[0].getText()

# PRODUTO 3
planCotacao['B5'] = name3[0].getText()
planCotacao['C5'] = date.today()
planCotacao['D5'] = str("R$") + prc3[0].getText()


# ------------------------------


# SALVAMENTO E FECHAMENTO DA PLANILHA
file.save(path)