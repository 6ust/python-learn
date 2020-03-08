from openpyxl import Workbook
from openpyxl import load_workbook

# Inserção de dados em planilha
file_excel = Workbook()


#comando utilizar planilha padrao
plan1 = file_excel.active


plan1.title = "Contas"

#comando utilizar criar nova planilha
plan2 = file_excel.create_sheet("emails")


#adicionar dados a planilha
plan1['B2'] = "Nome"
plan1['C2'] = "Sobrenome"
plan1['D2'] = "Email"
plan1['E2'] = "Nascimento"

print(file_excel.sheetnames)

#Salvar Planilha
file_excel.save("account_plan_py.xlsx")

#Carregar Planilha Existente
path = "account_plan_py.xlsx"
file_exist = load_workbook(path)